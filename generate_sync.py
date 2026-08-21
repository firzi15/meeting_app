import openpyxl
import json
import os

wb = openpyxl.load_workbook(r'C:\Users\Guest-Firzi\Downloads\Employee_Meeting.xlsx')
ws1 = wb['Sheet1']
ws2 = wb['Manager & Kabag']

mgr_kabag = {}
for r in range(2, ws2.max_row + 1):
    nik = str(ws2.cell(r, 2).value).strip() if ws2.cell(r, 2).value else ''
    jab = str(ws2.cell(r, 4).value).strip() if ws2.cell(r, 4).value else ''
    if nik:
        mgr_kabag[nik] = jab

branch_map = {
    'HEAD OFFICE': 1,
    'BRANCH SURABAYA': 2,
    'REPOSITORY KARAWANG': 3,
    'BRANCH BANDUNG': 4,
    'BRANCH MAKASSAR': 5,
    'BRANCH SEMARANG': 6,
    'BRANCH KLATEN': 7,
    'BRANCH MEDAN': 8,
    'BRANCH MANADO': 9
}

branch_names = [
    (1, 'Jakarta'),
    (2, 'Surabaya'),
    (3, 'Karawang'),
    (4, 'Bandung'),
    (5, 'Makassar'),
    (6, 'Semarang'),
    (7, 'Klaten'),
    (8, 'Medan'),
    (9, 'Manado')
]

cur_branch_str = 'HEAD OFFICE'
employees = []

for r in range(4, ws1.max_row + 1):
    c1 = ws1.cell(r, 1).value
    c2 = ws1.cell(r, 2).value
    c3 = ws1.cell(r, 3).value
    c5 = ws1.cell(r, 5).value
    c8 = ws1.cell(r, 8).value
    
    if c1 and isinstance(c1, str) and (c1.startswith('HEAD OFFICE') or c1.startswith('REPOSITORY') or c1.startswith('BRANCH')):
        cur_branch_str = c1.strip()
        continue
        
    if c2 and c3:
        nik = str(c2).strip()
        name = str(c3).strip()
        dept = str(c5).strip() if c5 else 'Umum'
        jab = str(c8).strip() if c8 else ''
        
        if nik in mgr_kabag and mgr_kabag[nik]:
            jab = mgr_kabag[nik]
            
        group = 'Staff'
        jab_u = jab.upper()
        if 'MANAGER' in jab_u:
            group = 'Manager'
        elif 'KEPALA BAGIAN' in jab_u or 'SECTION HEAD' in jab_u or 'KABAG' in jab_u:
            group = 'Kepala Bagian (Kabag)'
            
        b_id = branch_map.get(cur_branch_str, 1)
        username = nik.lower()
        
        employees.append({
            'nik': nik,
            'name': name,
            'username': username,
            'jabatan': jab,
            'group_name': group,
            'division': dept,
            'branch_id': b_id
        })

with open(r'c:\Users\Guest-Firzi\Documents\Meeting_App_Production\auto_sync_employees.json', 'w', encoding='utf-8') as f:
    json.dump({'branches': branch_names, 'employees': employees}, f, ensure_ascii=False, indent=2)

php_content = """<?php
// auto_sync_employees.php
require_once __DIR__ . '/database.php';

try {
    $dataFile = __DIR__ . '/auto_sync_employees.json';
    if (!file_exists($dataFile)) {
        die("auto_sync_employees.json not found.");
    }
    $payload = json_decode(file_get_contents($dataFile), true);
    if (!$payload) {
        die("Invalid JSON data.");
    }

    // 1. Ensure branches exist
    foreach ($payload['branches'] as $br) {
        $stmt = $pdo->prepare("INSERT INTO branches (id, name) VALUES (?, ?) ON CONFLICT (id) DO UPDATE SET name = EXCLUDED.name");
        $stmt->execute([$br[0], $br[1]]);
    }
    // Update sequence
    $pdo->exec("SELECT setval('branches_id_seq', (SELECT COALESCE(MAX(id), 1) FROM branches))");

    // 2. Ensure default groups exist
    $pdo->exec("INSERT INTO employee_groups (name, description) VALUES 
        ('Manager', 'Kelompok Managerial & Kepala Divisi/Cabang'),
        ('Kepala Bagian (Kabag)', 'Kelompok Kepala Bagian / Section Head'),
        ('Staff', 'Kelompok Staff Karyawan')
    ON CONFLICT (name) DO NOTHING");

    // 3. Upsert Employees
    $hashed_pass = password_hash('password123', PASSWORD_BCRYPT);

    $stmt_check = $pdo->prepare("SELECT id, role FROM users WHERE username = ? OR (nik = ? AND nik != '')");
    $stmt_insert = $pdo->prepare("INSERT INTO users (nik, name, username, password, role, jabatan, group_name, division, branch_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)");
    $stmt_update = $pdo->prepare("UPDATE users SET nik = ?, name = ?, jabatan = ?, group_name = ?, division = ?, branch_id = ? WHERE id = ?");

    foreach ($payload['employees'] as $emp) {
        $stmt_check->execute([$emp['username'], $emp['nik']]);
        $existing = $stmt_check->fetch();
        if ($existing) {
            $stmt_update->execute([$emp['nik'], $emp['name'], $emp['jabatan'], $emp['group_name'], $emp['division'], $emp['branch_id'], $existing['id']]);
        } else {
            $role = 'user';
            $stmt_insert->execute([$emp['nik'], $emp['name'], $emp['username'], $hashed_pass, $role, $emp['jabatan'], $emp['group_name'], $emp['division'], $emp['branch_id']]);
        }
    }
    
    // Ensure admin user is superadmin
    $pdo->exec("UPDATE users SET role = 'superadmin' WHERE username = 'admin'");
    
    echo "SYNC_SUCCESS: " . count($payload['employees']) . " employees processed successfully.";
} catch (Exception $e) {
    echo "SYNC_ERROR: " . $e->getMessage();
}
"""

with open(r'c:\Users\Guest-Firzi\Documents\Meeting_App_Production\auto_sync_employees.php', 'w', encoding='utf-8') as f:
    f.write(php_content)

print("auto_sync_employees.php and auto_sync_employees.json generated successfully.")
